# program that analyze excel sheet
# open cmd and type pip install openpyxl
# importing the package and name it xl
import openpyxl as xl

workbook = xl.load_workbook("xl.xlsx")
sheet = workbook['Sheet1']
# the first cell in the xlsx
cell = sheet['a1']
same_cell = sheet.cell(1, 1)
print(cell.value)
print(same_cell.value)

# print all column 3
for row in range(1, sheet.max_row + 1):
    cell = sheet.cell(row, 3)
    print(cell.value)

# get all values from the xlsx file
for index, row in enumerate(sheet.iter_rows()):
    for cell in row:
        print(cell.value)
